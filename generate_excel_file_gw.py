import os
import csv
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from requests import exceptions, get, post
from datetime import datetime

analysis_file_name = "gw_analysis_info"
now = datetime.now().strftime("%Y-%m-%d")
analysis_header_csv = ["ip", "port", "ping_status", "panel_name", "panel_number", "temp_db_count", "cpu_usage",
                       "gateway_uptime",
                       "process_uptime", "gateway_version", 'current_first_date', 'current_last_date',
                       'current_cloud_date', 'energy_first_date', 'energy_last_date', 'energy_cloud_date',
                       'event_first_date', 'event_last_date', 'event_cloud_date', 'frequency_first_date',
                       'frequency_last_date', 'frequency_cloud_date', 'misc_first_date', 'misc_last_date',
                       'misc_cloud_date', 'pcmcount_first_date', 'pcmcount_last_date', 'pcmcount_cloud_date',
                       'power_first_date', 'power_last_date', 'power_cloud_date', 'temperature_first_date',
                       'temperature_last_date', 'temperature_cloud_date', 'voltage_first_date', 'voltage_last_date',
                       'voltage_cloud_date', 'acmode_first_date', 'acmode_last_date', 'acmode_cloud_date',
                       'acspeed_first_date', 'acspeed_last_date', 'acspeed_cloud_date', 'acstate_first_date',
                       'acstate_last_date', 'acstate_cloud_date', 'event_first_date', 'event_last_date',
                       'event_cloud_date', 'humidity_first_date', 'humidity_last_date', 'humidity_cloud_date',
                       'acswing_first_date', 'acswing_last_date', 'acswing_cloud_date', 'dust_first_date',
                       'dust_last_date', 'dust_cloud_date', 'error_first_date', 'error_last_date', 'error_cloud_date',
                       'pm10_first_date', 'pm10_last_date', 'pm10_cloud_date', 'pm1p0_first_date', 'pm1p0_last_date',
                       'pm1p0_cloud_date', 'pm2p5_first_date', 'pm2p5_last_date', 'pm2p5_cloud_date',
                       'speed_first_date', 'speed_last_date', 'speed_cloud_date']


def find_measuremt_dates(ip, panel_number):
    analysis_measuremts = ["current", "energy", "event", "frequency", "misc", "pcmcount", "power", "temperature",
                           "voltage",
                           "acmode", "acspeed", "acstate", "event", "humidity", "acswing", "dust", "error", "pm10",
                           "pm1p0",
                           "pm2p5", "speed"]
    from influxdb import InfluxDBClient
    import influxdb_client

    # Create an InfluxDB client
    client = InfluxDBClient(host=ip, port=8086, database='temp_data', username='admin',
                            password='admin123', timeout=2)

    influx2_client = influxdb_client.InfluxDBClient(
        url="http://192.168.1.11:8087",
        username="root",
        password="rootrootroot",
        token="FWpBQQIAPlK9ZCWBhmzblkCsnGk87_FcwmY7-df8vF9eVdoNyscCkOA5UsRfZNqgx62rzW05v4TdbSmG3vyMFw==",
        org="iam"
    )

    mes_dict = {}
    try:
        if client.ping():
            for mes in analysis_measuremts:
                query = f'SELECT * FROM {mes} ORDER BY time ASC LIMIT 1'
                result = client.query(query)
                if len(result.raw['series']) == 0:
                    mes_dict.update({f"{mes}_first_date": 'Not Available'})
                    mes_dict.update({f"{mes}_last_date": 'Not Available'})
                else:
                    for point in result.get_points():
                        time = point['time'].replace('T', ' ').replace('Z', '')
                        mes_dict.update({f"{mes}_first_date": time})
                        # print(f"{mes}-First-Time:", time)
                    query1 = f'SELECT * FROM {mes} ORDER BY time DESC LIMIT 1'
                    result1 = client.query(query1)
                    for point in result1.get_points():
                        time = point['time'].replace('T', ' ').replace('Z', '')
                        mes_dict.update({f"{mes}_last_date": time})
                        # print(f"{mes}-Last-Time:", time)
                cloud_query = f'''
                            from(bucket: "data")
                            |> range(start: -45d, stop: now())
                            |> filter(fn: (r) => r["_measurement"] == "{mes}")
                            |> filter(fn: (r) => r["panel_no"] == "{panel_number}")
                            |> group(columns: ["panel_no"])
                            |> aggregateWindow(every: 60m, fn: last, createEmpty: false)
                            |> last()
                            |> keep(columns: ["_time"])
                        '''
                query_api = influx2_client.query_api()
                cloud_result = query_api.query(cloud_query)
                if len(cloud_result) == 0:
                    mes_dict.update({f"{mes}_cloud_date": 'Not Available'})
                else:
                    from datetime import timedelta
                    for table in cloud_result:
                        for record in table.records:
                            time = str(record.values['_time'] + timedelta(minutes=330))
                            mes_dict.update({f"{mes}_cloud_date": time})
            influx2_client.close()
            client.close()
            return mes_dict
        else:
            pass
    except Exception as e:
        for mes in analysis_measuremts:
            mes_dict.update({f"{mes}_first_date": 'Not Available'})
            mes_dict.update({f"{mes}_last_date": 'Not Available'})
            mes_dict.update({f"{mes}_cloud_date": 'Not Available'})
        return mes_dict


def analysis():
    try:
        temp_data = []
        file_path = f'/home/smartiam/PycharmProjects/Schedule-email-for-gateway/upload/all_gw_ip_list.csv'
        # file_path = typer.prompt("Enter the path of the file to import", type=str)
        record_count = 0
        with open(file_path, 'r') as csvfile:
            csv_reader = csv.reader(csvfile)
            # Skip the header row if it exists
            next(csv_reader, None)
            for row in csv_reader:
                record_count += 1
        logger.info(
            f"File was uploaded successfully Please wait {(record_count * 12) / 60} minutes while processing data")
        with open(file_path, mode='r') as file:
            reader = csv.DictReader(file)
            for lines in reader:
                temp_dict = {}
                ip = lines['ip']
                port = lines['port']
                panel_name = "Not Available"
                panel_number = "Not Available"
                cpu_usage = "Not Available"
                gw_uptime = "Not Available"
                gw_process_uptime = "Not Available"
                gw_version = "Not Available"
                # local_count = 0
                temp_count = 0
                ip_port = f"{ip}:{port}"
                try:
                    import datetime
                    ping_data = post(f'http://{ip_port}/ping', timeout=1)
                    # Here we check ping status
                    if ping_data.status_code == 200:
                        get_gw_status_details = get(f'http://{ip_port}/gateway_status/main_status')
                        # Here we get gw_status
                        if get_gw_status_details.status_code == 200:
                            cpu_usage = f"{get_gw_status_details.json().get('cpu_usage')}%"
                            uptime = get_gw_status_details.json().get('uptime')
                            gw_uptime = str(datetime.timedelta(seconds=uptime))
                            process_uptime = get_gw_status_details.json().get('process_uptime')
                            gw_process_uptime = str(datetime.timedelta(seconds=process_uptime))
                            gw_version = get_gw_status_details.json().get('version_no')

                            site_name_res = get(f"http://{ip_port}/gateway_detail")
                            # Here we get gw_details
                            if site_name_res.status_code == 200:
                                panel_name = site_name_res.json().get('panel_name')
                                panel_number = site_name_res.json().get('a_panel_no')

                                get_data = get(f'http://{ip_port}/influx/transactions')
                                # Here we get temp_db txn count
                                if get_data.status_code == 200:
                                    json_list = get_data.json()
                                    # local_count = sum(json_list['data'].values())
                                    rule_code = json_list['temp_data']['rule_code']
                                    if rule_code != 0:
                                        temp_count = sum(json_list['temp_data'].values()) - rule_code

                                final_mes_dict = find_measuremt_dates(ip, panel_number)
                                temp_dict.update(
                                    {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                                     "temp_db_count": temp_count, "cpu_usage": cpu_usage, "gateway_uptime": gw_uptime,
                                     "process_uptime": gw_process_uptime, "ping_status": "True",
                                     "gateway_version": gw_version})
                                temp_dict.update(final_mes_dict)
                                temp_data.append(temp_dict)
                            else:
                                pass
                        else:
                            pass
                    else:
                        pass
                except exceptions.ConnectionError or exceptions.ConnectTimeout as e:
                    final_mes_dict = find_measuremt_dates(ip, panel_number)
                    temp_dict.update(
                        {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                         "temp_db_count": temp_count, "cpu_usage": cpu_usage, "gateway_uptime": gw_uptime,
                         "process_uptime": gw_process_uptime, "ping_status": "False",
                         "gateway_version": gw_version})
                    temp_dict.update(final_mes_dict)
                    temp_data.append(temp_dict)
                except KeyError as r_code:
                    # print(r_code)
                    temp_count = sum(json_list['temp_data'].values())
                    final_mes_dict = find_measuremt_dates(ip, panel_number)
                    temp_dict.update(
                        {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                         "temp_db_count": temp_count, "cpu_usage": cpu_usage, "gateway_uptime": gw_uptime,
                         "process_uptime": gw_process_uptime, "ping_status": "True",
                         "gateway_version": gw_version})
                    temp_dict.update(final_mes_dict)
                    temp_data.append(temp_dict)

                except Exception as e:
                    # print(e)
                    final_mes_dict = find_measuremt_dates(ip, panel_number)
                    temp_dict.update(
                        {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                         "temp_db_count": temp_count, "cpu_usage": cpu_usage, "gateway_uptime": gw_uptime,
                         "process_uptime": gw_process_uptime, "ping_status": "False",
                         "gateway_version": gw_version})
                    temp_dict.update(final_mes_dict)
                    temp_data.append(temp_dict)
            # Sort temp_data in descending order based on temp_db_count
            temp_data.sort(key=lambda x: x['temp_db_count'], reverse=True)

            # Create a new Excel workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Set the header row
            for col_num, header in enumerate(analysis_header_csv, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

            # Define colors for red and green
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

            working_gw = 0
            not_working_gw = 0
            temp_data_gw = 0
            # Write data rows to the worksheet
            for row_num, row_data in enumerate(temp_data, 2):
                color_flag = False
                for col_num, header in enumerate(analysis_header_csv, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))

                    # Check if the header corresponds to a cloud date
                    from datetime import datetime, timedelta
                    from dateutil.parser import parse
                    from dateutil.tz import tzutc
                    today = datetime.now(tz=tzutc()) + timedelta(minutes=330)

                    # Check if the header corresponds to a cloud date
                    if header.endswith('_cloud_date'):
                        cloud_date_str = row_data.get(header)
                        if cloud_date_str != 'Not Available':
                            # Parse the date using dateutil.parser
                            cloud_date = parse(cloud_date_str)
                            # Calculate the difference in days
                            days_difference = (today - cloud_date).days
                            if 0 < days_difference < 2:
                                cell.fill = yellow_fill
                                # Apply red fill to the panel_number cell
                                panel_number = row_data.get('panel_number', '')
                                if panel_number:
                                    cell = ws.cell(row=row_num,
                                                   column=analysis_header_csv.index('panel_number') + 1)
                                    cell.fill = yellow_fill

                            if days_difference > 2:
                                color_flag = True
                                cell.fill = red_fill
                    if color_flag:
                        # Apply red fill to the panel_number cell
                        panel_number = row_data.get('panel_number', '')
                        if panel_number:
                            cell = ws.cell(row=row_num,
                                           column=analysis_header_csv.index('panel_number') + 1)
                            cell.fill = red_fill

                    # Check if the 'ping_status' is False, apply red fill, otherwise green fill
                    if header == 'ping_status' and row_data.get(header) == "False":
                        cell.fill = red_fill
                        not_working_gw += 1
                    elif header == 'ping_status' and row_data.get(header) == "True":
                        cell.fill = green_fill
                        working_gw += 1
                if row_data.get('temp_db_count', 0) > 1500:
                    temp_data_gw += 1

            start_row = len(temp_data) + 3

            # Dictionary with labels and corresponding values
            label_value_dict = {
                "Working Gateways": (working_gw - temp_data_gw),
                "Not Working Gateways": not_working_gw,
                "TempDB Gateways": temp_data_gw
            }

            # Loop through the dictionary and set them in the first and second columns
            for label, value in label_value_dict.items():
                ws.cell(row=start_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=start_row, column=2, value=value)
                start_row += 1  # Move to the next row

            download_path = os.path.join('/home/smartiam/PycharmProjects/Schedule-email-for-gateway', 'download')
            if not os.path.exists(download_path):
                os.makedirs(download_path)
            excel_file_name = f'{analysis_file_name}_{now}.xlsx'
            excel_file_path = os.path.join(download_path, excel_file_name)
            wb.save(excel_file_path)
            logger.info(f"File is created at {excel_file_path}")
    except Exception as e:
        logger.error(str(e))
        logger.error("UNSUCCESSFUL")
