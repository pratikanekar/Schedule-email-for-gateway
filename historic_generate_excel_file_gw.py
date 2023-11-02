import os
import csv
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from requests import exceptions, get, post
from datetime import datetime

historic_file_name = "historic_analysis_report_gw"
now = datetime.now().strftime("%Y-%m-%d")
analysis_header_csv = ["ip", "port", "ping_status", "panel_name", "panel_number", "total_devices", "online_devices",
                       "local_db_count", "temp_db_count",
                       "last_HB_cloud_date", "last_DATA_cloud_date", "cpu_usage", "disk_usage", "memory_usage",
                       "gateway_uptime",
                       "process_uptime",
                       "gateway_version"]


def find_cloud_dates(panel_number):
    mes_dict = {}
    try:
        import influxdb_client
        import pandas as pd
        influx2_client = influxdb_client.InfluxDBClient(
            url="http://192.168.1.11:8087",
            username="root",
            password="rootrootroot",
            token="FWpBQQIAPlK9ZCWBhmzblkCsnGk87_FcwmY7-df8vF9eVdoNyscCkOA5UsRfZNqgx62rzW05v4TdbSmG3vyMFw==",
            org="iam"
        )
        hb_cloud_query = f'''
                    from(bucket: "data")
                    |> range(start: -90d, stop: now())
                    |> filter(fn: (r) => r["a_panel_number"] == "{panel_number}")
                    |> filter(fn: (r) => r["_measurement"] == "sys_info")
                    |> timeShift(duration: 330m, columns: ["_start", "_stop", "_time"])
                    |> aggregateWindow(every: 60m, fn: last, createEmpty: false)
                    |> filter(fn: (r) => r._field == "core_voltage")
                    |> last()
                    |> keep(columns: ["_time"])
                '''
        query_api = influx2_client.query_api()
        hb_cloud_result = query_api.query(hb_cloud_query)
        if len(hb_cloud_result) == 0:
            mes_dict.update({f"last_HB_cloud_date": 'Not Available'})
        else:
            for table in hb_cloud_result:
                for record in table.records:
                    time = str(record.values['_time'])
                    mes_dict.update({f"last_HB_cloud_date": time})
        result_list = []
        data_cloud_query = f'''
                            from(bucket: "data")
                            |> range(start: -90d, stop: now())
                            |> filter(fn: (r) => r["panel_no"] == "{panel_number}")
                            |> timeShift(duration: 330m, columns: ["_start", "_stop", "_time"])
                            |> last()
                        '''
        query_api = influx2_client.query_api()
        data_cloud_result = query_api.query(data_cloud_query)
        if len(data_cloud_result) == 0:
            mes_dict.update({f"last_DATA_cloud_date": 'Not Available'})
        else:
            for table in data_cloud_result:
                for record in table.records:
                    result_list.append(record.values)
            data_frame = pd.DataFrame(result_list)
            pd.set_option('display.max_columns', None)
            time = str(data_frame.loc[:, '_time'].max())
            mes_dict.update({f"last_DATA_cloud_date": time})
        influx2_client.close()
        return mes_dict
    except Exception as e:
        mes_dict.update({f"last_DATA_cloud_date": 'Not Available'})
        mes_dict.update({f"last_HB_cloud_date": 'Not Available'})
        return mes_dict


def historic_analysis():
    try:
        temp_data = []
        file_path = f'/home/smartiam/PycharmProjects/Schedule-email-for-gateway/upload/all_gw_ip_list.csv'
        # file_path = typer.prompt("Enter the path of the file to import", type=str)
        record_count = 0
        with open(file_path, 'r') as csvfile:
            csv_reader = csv.reader(csvfile)
            # Skip the header row if it exists
            next(csv_reader, None)
            for row in csv_reader:
                record_count += 1
        logger.info(
            f"File was uploaded successfully Please wait approximately {round((record_count * 5) / 60, 2)} minutes while processing historic data")
        with open(file_path, mode='r') as file:
            reader = csv.DictReader(file)
            for lines in reader:
                temp_dict = {}
                ip = lines['ip']
                port = lines['port']
                panel_name = lines['panel_name']
                panel_number = lines['panel_number']
                cpu_usage = "-"
                disk_usage = "-"
                memory_usage = "-"
                gw_uptime = "-"
                gw_process_uptime = "-"
                gw_version = "-"
                total_dev = 0
                online_dev = 0
                local_count = 0
                temp_count = 0
                ip_port = f"{ip}:{port}"
                try:
                    import datetime
                    ping_data = post(f'http://{ip_port}/ping', timeout=2)
                    # Here we check ping status
                    if ping_data.status_code == 200:
                        get_gw_details = get(f'http://{ip_port}/gateway_detail')
                        # Here we get gw_details for panel info
                        if get_gw_details.status_code == 200:
                            panel_number = get_gw_details.json().get('a_panel_no')
                            panel_name = get_gw_details.json().get('panel_name')
                        else:
                            pass
                        get_status_code_details = get(f'http://{ip_port}/status_codes')
                        if get_status_code_details.status_code == 200:
                            total_dev = len(get_status_code_details.json().get('devices'))
                            device_list = get_status_code_details.json().get('devices')
                            for key, value in device_list.items():
                                current_status = value['current']
                                if int(current_status) < 2000:
                                    online_dev += 1
                        else:
                            pass
                        get_gw_status_details = get(f'http://{ip_port}/gateway_status/main_status')
                        # Here we get gw_status
                        if get_gw_status_details.status_code == 200:
                            cpu_usage = f"{get_gw_status_details.json().get('cpu_usage')}%"
                            disk_usage = f"{get_gw_status_details.json().get('disk_usage')}%"
                            memory_usage = f"{get_gw_status_details.json().get('memory_usage')}%"
                            uptime = get_gw_status_details.json().get('uptime')
                            gw_uptime = str(datetime.timedelta(seconds=uptime))
                            process_uptime = get_gw_status_details.json().get('process_uptime')
                            gw_process_uptime = str(datetime.timedelta(seconds=process_uptime))
                            gw_version = get_gw_status_details.json().get('version_no')

                            get_data = get(f'http://{ip_port}/influx/transactions')
                            # Here we get temp_db txn count
                            if get_data.status_code == 200:
                                json_list = get_data.json()
                                local_count = sum(json_list['data'].values())
                                rule_code = json_list['temp_data']['rule_code']
                                if rule_code != 0:
                                    temp_count = sum(json_list['temp_data'].values()) - rule_code

                            final_mes_dict = find_cloud_dates(panel_number)
                            temp_dict.update(
                                {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                                 "total_devices": total_dev, "online_devices": online_dev,
                                 "local_db_count": local_count, "temp_db_count": temp_count, "cpu_usage": cpu_usage,
                                 "disk_usage": disk_usage, "memory_usage": memory_usage, "gateway_uptime": gw_uptime,
                                 "process_uptime": gw_process_uptime, "ping_status": "True",
                                 "gateway_version": gw_version})
                            temp_dict.update(final_mes_dict)
                            temp_data.append(temp_dict)

                        else:
                            pass
                    else:
                        final_mes_dict = find_cloud_dates(panel_number)
                        temp_dict.update(
                            {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                             "total_devices": total_dev, "online_devices": online_dev,
                             "local_db_count": local_count, "temp_db_count": temp_count, "cpu_usage": cpu_usage,
                             "disk_usage": disk_usage, "memory_usage": memory_usage, "gateway_uptime": gw_uptime,
                             "process_uptime": gw_process_uptime, "ping_status": "True",
                             "gateway_version": gw_version})
                        temp_dict.update(final_mes_dict)
                        temp_data.append(temp_dict)
                except exceptions.ConnectionError or exceptions.ConnectTimeout as e:
                    final_mes_dict = find_cloud_dates(panel_number)
                    temp_dict.update(
                        {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                         "total_devices": total_dev, "online_devices": online_dev,
                         "local_db_count": local_count, "temp_db_count": temp_count, "cpu_usage": cpu_usage,
                         "disk_usage": disk_usage, "memory_usage": memory_usage, "gateway_uptime": gw_uptime,
                         "process_uptime": gw_process_uptime, "ping_status": "True",
                         "gateway_version": gw_version})
                    temp_dict.update(final_mes_dict)
                    temp_data.append(temp_dict)
                except KeyError as r_code:
                    # print(r_code)
                    temp_count = sum(json_list['temp_data'].values())
                    final_mes_dict = find_cloud_dates(panel_number)
                    temp_dict.update(
                        {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                         "total_devices": total_dev, "online_devices": online_dev,
                         "local_db_count": local_count, "temp_db_count": temp_count, "cpu_usage": cpu_usage,
                         "disk_usage": disk_usage, "memory_usage": memory_usage, "gateway_uptime": gw_uptime,
                         "process_uptime": gw_process_uptime, "ping_status": "True",
                         "gateway_version": gw_version})
                    temp_dict.update(final_mes_dict)
                    temp_data.append(temp_dict)

                except Exception as e:
                    # print(e)
                    final_mes_dict = find_cloud_dates(panel_number)
                    temp_dict.update(
                        {"ip": ip, "port": port, "panel_name": panel_name, "panel_number": panel_number,
                         "total_devices": total_dev, "online_devices": online_dev,
                         "local_db_count": local_count, "temp_db_count": temp_count, "cpu_usage": cpu_usage,
                         "disk_usage": disk_usage, "memory_usage": memory_usage, "gateway_uptime": gw_uptime,
                         "process_uptime": gw_process_uptime, "ping_status": "True",
                         "gateway_version": gw_version})
                    temp_dict.update(final_mes_dict)
                    temp_data.append(temp_dict)
            # Sort temp_data in descending order based on temp_db_count
            temp_data.sort(key=lambda x: x['temp_db_count'], reverse=True)

            # Create a new Excel workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Set the header row
            for col_num, header in enumerate(analysis_header_csv, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)

            # Define colors for red and green
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

            working_gw = 0
            not_working_gw = 0
            temp_data_gw = 0
            # Write data rows to the worksheet
            for row_num, row_data in enumerate(temp_data, 2):
                color_flag = False
                row_color_flag = False
                for col_num, header in enumerate(analysis_header_csv, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))

                    # Check if the header corresponds to a cloud date
                    from datetime import datetime, timedelta
                    from dateutil.parser import parse
                    from dateutil.tz import tzutc
                    today = datetime.now(tz=tzutc()) + timedelta(minutes=330)

                    # Check if the header corresponds to a cloud date
                    if header.endswith('_cloud_date'):
                        cloud_date_str = row_data.get(header)
                        if cloud_date_str != 'Not Available':
                            # Parse the date using dateutil.parser
                            cloud_date = parse(cloud_date_str)
                            # Calculate the difference in days
                            days_difference = (today - cloud_date).days
                            if days_difference > 10:
                                row_color_flag = True

                            if 0 < days_difference < 2:
                                cell.fill = yellow_fill
                                # Apply red fill to the panel_number cell
                                panel_number = row_data.get('panel_number', '')
                                if panel_number:
                                    cell = ws.cell(row=row_num,
                                                   column=analysis_header_csv.index('panel_number') + 1)
                                    cell.fill = yellow_fill

                            if days_difference > 2:
                                color_flag = True
                                cell.fill = red_fill
                        else:
                            # Set the entire row to red
                            for col_num in range(1, len(analysis_header_csv) + 2):
                                cell = ws.cell(row=row_num, column=col_num)
                                cell.fill = red_fill

                    if row_color_flag:
                        # Set the entire row to red
                        for col_num in range(1, len(analysis_header_csv) + 2):
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.fill = red_fill
                    if color_flag:
                        # Apply red fill to the panel_number cell
                        panel_number = row_data.get('panel_number', '')
                        if panel_number:
                            cell = ws.cell(row=row_num,
                                           column=analysis_header_csv.index('panel_number') + 1)
                            cell.fill = red_fill

                    # Check if the 'ping_status' is False, apply red fill, otherwise green fill
                    if header == 'ping_status' and row_data.get(header) == "False":
                        cell.fill = red_fill
                        not_working_gw += 1
                    elif header == 'ping_status' and row_data.get(header) == "True":
                        cell.fill = green_fill
                        working_gw += 1
                if row_data.get('temp_db_count', 0) > 1500:
                    temp_data_gw += 1

            start_row = len(temp_data) + 3

            # Dictionary with labels and corresponding values
            label_value_dict = {
                "Working Gateways": (working_gw - temp_data_gw),
                "Not Working Gateways": not_working_gw,
                "TempDB Gateways": temp_data_gw
            }

            # Loop through the dictionary and set them in the first and second columns
            for label, value in label_value_dict.items():
                ws.cell(row=start_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=start_row, column=2, value=value)
                start_row += 1  # Move to the next row

            download_path = os.path.join('/home/smartiam/PycharmProjects/Schedule-email-for-gateway',
                                         'historic_report_download')
            if not os.path.exists(download_path):
                os.makedirs(download_path)
            excel_file_name = f'{historic_file_name}_{now}.xlsx'
            excel_file_path = os.path.join(download_path, excel_file_name)
            wb.save(excel_file_path)
            logger.info(f"File is created at {excel_file_path}")
    except Exception as e:
        logger.error(str(e))
        logger.error("UNSUCCESSFUL")
